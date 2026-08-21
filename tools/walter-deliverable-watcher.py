#!/usr/bin/env python3
"""walter-deliverable-watcher.py — Server-seitiger Hook für Walter-Deliverable-Mails.

Pollt regelmäßig die Paperclip-API, findet neu auf `done` gegangene Walter-Top-Level-Issues
(parentId=null, createdByUserId=WALTER) und schickt eine Komplettpaket-Mail an Walter:
- Auftrag (issue.description)
- Sub-Issue-Übersicht mit Status und Last-Comment-Summary
- Finale Zusammenfassung des Root-Issues
- Alle erzeugten Vault-MDs inline (im Body) UND als Anhang

State: ~/.paperclip/instances/default/state/walter-deliverable-sent.json
Log:   ~/.paperclip/instances/default/logs/walter-deliverable-watcher.log

Designed-Spec siehe Chatverlauf 2026-05-15 mit Walter.
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
import urllib.request
import urllib.error
import subprocess
from datetime import datetime, timezone
from pathlib import Path

# --- Konstanten ---------------------------------------------------------------

COMPANIES = [
  {
    "id": "9cebf3cf-efe8-4597-a400-f06488900a87",  # WHITESTAG
    "issue_url_prefix": "https://company.whitestag.ai/WHI/issues/",
    "creator_filter": None,  # set below after WALTER_USER_ID is defined
    "identifier_allowlist": None,
  },
  {
    "id": "158c4959-4973-4cb0-8066-55ec0f35625e",  # Health Insights
    "issue_url_prefix": "https://company.whitestag.ai/HEA/issues/",
    "creator_filter": None,  # HEA-94 wird vom Heartbeat erzeugt
    "identifier_allowlist": {"HEA-94"},
  },
]
PAPERCLIP_BASE = os.environ.get("PAPERCLIP_API_URL", "http://127.0.0.1:3100").rstrip("/")
API_BASE = f"{PAPERCLIP_BASE}/api"
WEBHOOK_URL = "http://127.0.0.1:5678/webhook/mailhub/send"
def _load_mailhub_secret() -> str:
    """Secret aus der zentralen Secrets-Datei lesen (Rotation 03.08.2026, nicht mehr im Code)."""
    path = os.path.expanduser("~/.paperclip/instances/default/secrets/mailhub.env")
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            if line.startswith("MAILHUB_SECRET="):
                return line.split("=", 1)[1].strip().strip('"')
    raise RuntimeError("MAILHUB_SECRET nicht gefunden in " + path)


MAILHUB_SECRET = _load_mailhub_secret()
WALTER_USER_ID = "18r34Ghx5N0LHRptMCT6Fp1WaoGqhvc9"
COMPANIES[0]["creator_filter"] = WALTER_USER_ID
TO_ADDR = "ws@whitestag.ai"
FROM_ADDR = "office@whitestag.ai"
VAULT_ROOT = "/Users/walterschoenenbroecher.de/Obsidian/WHITESTAG-Vault"

HOME = os.path.expanduser("~")
AUTH_FILE = f"{HOME}/.paperclip/auth.json"
STATE_FILE = f"{HOME}/.paperclip/instances/default/state/walter-deliverable-sent.json"
LOG_FILE = f"{HOME}/.paperclip/instances/default/logs/walter-deliverable-watcher.log"

POLL_INTERVAL_S = 60
LOOKBACK_LIMIT = 500
MAX_ATTACH_TOTAL_BYTES = 24 * 1024 * 1024
MAX_INLINE_FILE_BYTES = 200 * 1024  # >200KB nur als Anhang, nicht inline einbetten

# Deliverable-Formate: .md wird inline eingebettet, die Office-/PDF-Formate
# nur als Anhang (siehe select_inline_attachments).
DELIVERABLE_EXT = (".md", ".docx", ".xlsx", ".pdf")
INLINE_EXT = (".md",)

_VAULT_EXT_GROUP = "|".join(e.lstrip(".") for e in DELIVERABLE_EXT)
VAULT_PATH_RE = re.compile(
    r"(/Users/walterschoenenbroecher.de/Obsidian/WHITESTAG-Vault/[^\s\"'`)>\]]+\.(?:"
    + _VAULT_EXT_GROUP
    + r"))"
)
VAULT_DIR_RE = re.compile(
    r"(/Users/walterschoenenbroecher.de/Obsidian/WHITESTAG-Vault/[^\s\"'`)>\]]+?)(?=[\s\"'`)>\]]|$)"
)
# Office-/PDF-Deliverables liegen oft außerhalb des Vaults (z.B. im Projekt-
# Ordner Dokumente/). Diese werden per absolutem Pfad in Issue-Texten zitiert
# und über os.path.isfile validiert — .md bleibt bewusst vault-beschränkt,
# um keine fremden Notizen einzusammeln.
# Leerzeichen sind erlaubt (Word-Deliverables heißen z.B. "Angebot Kunde V2.docx"),
# non-greedy bis zur ersten Endung; gestoppt wird an Quote/Backtick/Klammer/Zeilen-
# umbruch. Falsch zusammengezogene Pfade fängt der os.path.isfile-Guard ab.
ABS_DOC_RE = re.compile(r"(/[^\r\n\"'`)>\]<|]+?\.(?:docx|xlsx|pdf))")

# Mail-Spiegel, Heartbeat- und Log-Verzeichnisse sind KEINE Deliverables.
# Die E-Mails/*.md sind n8n-Spiegel ein-/ausgehender Mails und tragen oft den
# Issue-Identifier im Namen (z.B. "2026-06-21-[WHI-1662]-…-ceo.md"). Ohne diesen
# Filter zieht discover_vault_files (Pfad 2: Identifier-Match über den ganzen
# Vault) sie ein und bettet damit alte Mails rekursiv in neue Deliverable-Mails.
EXCLUDED_VAULT_SEGMENTS = ("/E-Mails/", "/Heartbeat/", "/Logs/")
EXCLUDED_VAULT_BASENAMES = ("Mailhub-Outbound-Log.md",)


def _is_excluded_vault_file(path: str) -> bool:
    if any(seg in path for seg in EXCLUDED_VAULT_SEGMENTS):
        return True
    return os.path.basename(path) in EXCLUDED_VAULT_BASENAMES


# --- Helper -------------------------------------------------------------------


LOG_MAX_BYTES = 2 * 1024 * 1024  # ab 2 MB rotieren (.1 behalten)


def _rotate_if_large(path: str) -> None:
    try:
        if os.path.getsize(path) > LOG_MAX_BYTES:
            os.replace(path, path + ".1")
    except OSError:
        pass


def log(level: str, msg: str) -> None:
    os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
    _rotate_if_large(LOG_FILE)
    line = f"{datetime.now(timezone.utc).isoformat()}\t{level}\t{msg}\n"
    with open(LOG_FILE, "a") as f:
        f.write(line)
    if level in ("ERROR", "WARN"):
        sys.stderr.write(line)


def load_token() -> str:
    with open(AUTH_FILE) as f:
        data = json.load(f)
    creds = data["credentials"]
    # auth.json ist nach der Ausstellungs-URL geschluesselt: erst die
    # konfigurierte Adresse, dann die historischen Schreibweisen, zuletzt
    # der einzige Eintrag.
    for _key in (PAPERCLIP_BASE, "http://localhost:3100", "http://127.0.0.1:3100"):
        if _key in creds:
            return creds[_key]["token"]
    if len(creds) == 1:
        return next(iter(creds.values()))["token"]
    raise KeyError(f"Kein Token fuer {PAPERCLIP_BASE} in der auth.json")


class AuthError(Exception):
    """Raised when the Paperclip API rejects our board token (401/403).

    Board-Tokens haben eine TTL von 30 Tagen; danach kommt 401 und der Watcher
    muss laut alarmieren statt still in eine Crash-Schleife zu laufen.
    """


def api_get(token: str, path: str) -> object:
    req = urllib.request.Request(API_BASE + path, headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        if e.code in (401, 403):
            raise AuthError(f"API {e.code} on {path} — board token expired/invalid") from e
        raise


def load_state() -> dict:
    os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
    if not os.path.exists(STATE_FILE):
        return {"sent": {}}
    try:
        with open(STATE_FILE) as f:
            return json.load(f)
    except json.JSONDecodeError:
        log("WARN", f"state file corrupt, resetting: {STATE_FILE}")
        return {"sent": {}}


def save_state(state: dict) -> None:
    tmp = STATE_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump(state, f, indent=2, sort_keys=True)
    os.replace(tmp, STATE_FILE)


# --- Tree-Walk ---------------------------------------------------------------


def fetch_full_tree(token: str, root_id: str, company_id: str) -> dict:
    """Sammelt root + alle Sub-Issues rekursiv. Returns {id: issue_dict}."""
    tree: dict = {}
    queue = [root_id]
    while queue:
        nxt = queue.pop()
        if nxt in tree:
            continue
        try:
            issue = api_get(token, f"/issues/{nxt}")
        except Exception as e:
            log("WARN", f"tree-fetch failed for {nxt}: {e}")
            continue
        tree[nxt] = issue
        try:
            children = api_get(
                token, f"/companies/{company_id}/issues?parentId={nxt}"
            )
            items = children if isinstance(children, list) else children.get("issues", [])
            for c in items:
                if isinstance(c, dict) and c.get("id"):
                    queue.append(c["id"])
        except Exception as e:
            log("WARN", f"children-fetch failed for {nxt}: {e}")
    return tree


def fetch_comments(token: str, issue_id: str) -> list[dict]:
    try:
        d = api_get(token, f"/issues/{issue_id}/comments")
        items = d if isinstance(d, list) else d.get("comments", [])
        return [c for c in items if isinstance(c, dict)]
    except Exception as e:
        log("WARN", f"comments fetch failed for {issue_id}: {e}")
        return []


# --- Vault-MD-Erkennung -------------------------------------------------------


def _parse_iso(s: str | None) -> float | None:
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).timestamp()
    except Exception:
        return None


def discover_vault_files(tree: dict, comments_by_issue: dict[str, list[dict]]) -> list[str]:
    found: set[str] = set()
    text_corpus: list[str] = []

    # 1) Vault-Pfade in Beschreibungen + Kommentaren (exact hits, alle Formate)
    #    plus absolute Office-/PDF-Pfade außerhalb des Vaults (z.B. Dokumente/).
    def scan(text: str | None) -> None:
        if not text:
            return
        text_corpus.append(text)
        for m in VAULT_PATH_RE.findall(text):
            p = m.rstrip(".,;:!?")
            if os.path.isfile(p):
                found.add(os.path.realpath(p))
        for m in ABS_DOC_RE.findall(text):
            p = m.rstrip(".,;:!?")
            if os.path.isfile(p):
                found.add(os.path.realpath(p))

    for issue in tree.values():
        scan(issue.get("description"))
        scan(issue.get("title"))
    for cmts in comments_by_issue.values():
        for c in cmts:
            scan(c.get("body"))

    # 2) Vault-Files mit Issue-Identifier im Dateinamen (z.B. "WHI-454 ...md")
    identifiers = {i.get("identifier") for i in tree.values() if i.get("identifier")}
    if identifiers:
        for dirpath, _, filenames in os.walk(VAULT_ROOT):
            if "/.obsidian" in dirpath or "/.trash" in dirpath:
                continue
            for fn in filenames:
                if not fn.endswith(DELIVERABLE_EXT):
                    continue
                for ident in identifiers:
                    if ident in fn:
                        found.add(os.path.realpath(os.path.join(dirpath, fn)))
                        break

    # 3) Vault-Directories aus Issue-Texten extrahieren und MDs im Zeitfenster nehmen.
    #    Issue erwähnt z.B. "Vault-Pfad für diese Domain: /Volumes/.../claude-api"
    #    → wir nehmen alle .md in diesem Directory, deren mtime im Issue-Lifecycle liegt.
    starts = [_parse_iso(i.get("createdAt")) for i in tree.values()]
    ends = [_parse_iso(i.get("completedAt")) for i in tree.values()]
    starts = [s for s in starts if s]
    ends = [e for e in ends if e]
    if starts and ends:
        win_start = min(starts) - 60  # 1 min Puffer
        win_end = max(ends) + 5 * 60  # 5 min Puffer
        dirs: set[str] = set()
        joined = "\n".join(text_corpus)
        for m in VAULT_DIR_RE.findall(joined):
            p = m.rstrip(".,;:!?")
            if p.endswith(".md"):
                continue
            if os.path.isdir(p):
                dirs.add(os.path.realpath(p))
        for d in dirs:
            try:
                with os.scandir(d) as it:
                    for entry in it:
                        if not entry.is_file() or not entry.name.endswith(DELIVERABLE_EXT):
                            continue
                        try:
                            mt = entry.stat().st_mtime
                        except OSError:
                            continue
                        if win_start <= mt <= win_end:
                            found.add(os.path.realpath(entry.path))
            except OSError:
                continue

    return sorted(p for p in found if not _is_excluded_vault_file(p))


# --- Summary-Extraktion -------------------------------------------------------


def is_system_comment(c: dict) -> bool:
    body = (c.get("body") or "").strip()
    if not body:
        return True
    starters = (
        "Paperclip needs a disposition",
        "Paperclip automatically retried",
        "Task ",  # "Task WHI-XXX marked as done."
        "**Adapter post-run guard auto-closed",
        "Delegiert an ",  # generischer Delegations-Marker, oft inhaltsleer
    )
    if any(body.startswith(s) for s in starters):
        return len(body) < 200  # nur wenn kurz, sonst evtl. nützlich
    return False


def extract_summary(comments: list[dict]) -> str:
    # Letzter substantieller, nicht-System-Kommentar
    for c in sorted(comments, key=lambda x: x.get("createdAt") or "", reverse=True):
        body = (c.get("body") or "").strip()
        if len(body) < 30:
            continue
        if is_system_comment(c):
            continue
        return body
    # Fallback: jeglicher Kommentar
    for c in sorted(comments, key=lambda x: x.get("createdAt") or "", reverse=True):
        body = (c.get("body") or "").strip()
        if body:
            return body
    return "_(keine Zusammenfassung im Issue)_"


# --- Mail-Build ---------------------------------------------------------------


def fmt_dt(iso: str | None) -> str:
    if not iso:
        return "?"
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return iso


def build_mail(
    root: dict,
    tree: dict,
    comments_by_issue: dict[str, list[dict]],
    vault_files: list[str],
    inline_files: list[str],
) -> tuple[str, str, str]:
    """Returns (subject, text_body, html_body) — text is the plaintext fallback,
    html is rendered via the shared lib/build_walter_mail_html.py helper."""
    ident = root.get("identifier") or root.get("id")
    title = root.get("title") or "(ohne Titel)"
    subject = f"[{ident}] {title} — Auftrag abgeschlossen"

    # --- Sektion: Aufgabe (root.description) ---
    aufgabe_md = (root.get("description") or "_(keine Beschreibung)_").strip()

    # --- Sektion: Zusammenfassung (finale Zusammenfassung des Root-Issues) ---
    zusammenfassung_md = extract_summary(comments_by_issue.get(root.get("id"), []))

    # --- Sektion: Ergebnis (Sub-Issue-Outputs + inline-Dokumente) ---
    ergebnis_lines: list[str] = []
    subs = [
        i for i in tree.values()
        if i.get("id") != root.get("id") and i.get("parentId") == root.get("id")
    ]
    if subs:
        ergebnis_lines.append("### Sub-Issues (Tiefe 1)")
        ergebnis_lines.append("")
        for s in sorted(subs, key=lambda x: x.get("identifier") or ""):
            s_ident = s.get("identifier")
            s_title = s.get("title")
            s_status = s.get("status")
            ergebnis_lines.append(f"#### {s_ident} — {s_title}")
            ergebnis_lines.append(f"Status: **{s_status}**")
            ergebnis_lines.append("")
            ergebnis_lines.append(extract_summary(comments_by_issue.get(s.get("id"), [])))
            ergebnis_lines.append("")
    if inline_files:
        if ergebnis_lines:
            ergebnis_lines.append("---")
            ergebnis_lines.append("")
        for p in inline_files:
            ergebnis_lines.append(f"### Dokument: {os.path.basename(p)}")
            ergebnis_lines.append("")
            try:
                with open(p) as f:
                    ergebnis_lines.append(f.read())
            except Exception as e:
                ergebnis_lines.append(f"_(Datei nicht lesbar: {e})_")
            ergebnis_lines.append("")
    if not ergebnis_lines:
        ergebnis_lines.append("_(Keine Deliverable-Datei erzeugt — Issue wurde ohne Vault-Output abgeschlossen.)_")
    ergebnis_md = "\n".join(ergebnis_lines)

    # --- Sektion: Restliche Informationen (Metadaten + alle Vault-Dateien) ---
    restliche_lines: list[str] = []
    restliche_lines.append("| Feld | Wert |")
    restliche_lines.append("|---|---|")
    company = root.get("_company") or COMPANIES[0]
    issue_url = f"{company['issue_url_prefix']}{ident}"
    restliche_lines.append(f"| Issue | [{ident} — {title}]({issue_url}) |")
    restliche_lines.append(f"| Status | `{root.get('status')}` |")
    restliche_lines.append(f"| Auftrag erteilt | {fmt_dt(root.get('createdAt'))} |")
    restliche_lines.append(f"| Abgeschlossen | {fmt_dt(root.get('completedAt'))} |")
    if vault_files:
        rels = [p.replace(VAULT_ROOT + "/", "") for p in vault_files]
        restliche_lines.append(f"| Dokumente | {'<br>'.join(f'`{r}`' for r in rels)} |")
    restliche_md = "\n".join(restliche_lines)

    # --- HTML via Shared Builder ---
    spec = {
        "title": title,
        "title_icon": "📋",
        "subtitle_html": (
            f"<a href='{issue_url}' style='color:#0066cc;text-decoration:none;'>{ident}</a>"
            f" · Auftrag abgeschlossen"
        ),
        "sections": [
            {"icon": "🎯", "title": "Aufgabe", "body_md": aufgabe_md, "panel": True},
            {"icon": "📝", "title": "Zusammenfassung", "body_md": zusammenfassung_md},
            {"icon": "✅", "title": "Ergebnis", "body_md": ergebnis_md},
            {"icon": "ℹ️", "title": "Restliche Informationen", "body_md": restliche_md},
        ],
    }
    html_body = _render_mail_html(spec, company["id"])

    # --- Plaintext-Fallback (klassisch, für nicht-HTML-Clients) ---
    text_lines: list[str] = []
    text_lines.append(f"# {title}")
    text_lines.append("")
    text_lines.append(f"Issue: {ident} ({issue_url})")
    text_lines.append(f"Status: {root.get('status')}")
    text_lines.append(f"Auftrag erteilt: {fmt_dt(root.get('createdAt'))}")
    text_lines.append(f"Abgeschlossen: {fmt_dt(root.get('completedAt'))}")
    text_lines.append("")
    text_lines.append("## AUFGABE")
    text_lines.append("")
    text_lines.append(aufgabe_md)
    text_lines.append("")
    text_lines.append("## ZUSAMMENFASSUNG")
    text_lines.append("")
    text_lines.append(zusammenfassung_md)
    text_lines.append("")
    text_lines.append("## ERGEBNIS")
    text_lines.append("")
    text_lines.append(ergebnis_md)
    text_lines.append("")
    if vault_files:
        text_lines.append("## DOKUMENTE")
        text_lines.append("")
        for p in vault_files:
            text_lines.append(f"- {p.replace(VAULT_ROOT + '/', '')}")
    text_body = "\n".join(text_lines)

    return subject, text_body, html_body


def _render_mail_html(spec: dict, company_id: str) -> str:
    """Render mail HTML via the shared lib/build_walter_mail_html.py helper.
    Imports the module lazily so this watcher script keeps working even if the
    helper is unavailable (then returns empty string → mail goes out text-only)."""
    import importlib.util
    helper_path = os.path.join(
        HOME,
        ".paperclip/instances/default/companies",
        company_id,
        "lib/build_walter_mail_html.py",
    )
    if not os.path.exists(helper_path):
        log("WARN", f"mail-html helper missing at {helper_path} — sending text-only")
        return ""
    try:
        spec_module = importlib.util.spec_from_file_location("build_walter_mail_html", helper_path)
        if not spec_module or not spec_module.loader:
            return ""
        mod = importlib.util.module_from_spec(spec_module)
        spec_module.loader.exec_module(mod)
        return mod.build_html(spec)
    except Exception as exc:  # noqa: BLE001
        log("WARN", f"mail-html helper failed: {exc} — sending text-only")
        return ""



# --- Lücke 1: .md → .docx Konvertierung --------------------------------------

DOCX_OUTPUT_DIR = os.path.join(VAULT_ROOT, "Paperclip/_Meta/docx-exports")


def count_words(text: str) -> int:
    """Zählt Wörter in einem Text-String."""
    if not text:
        return 0
    # Entferne Markdown-Syntax für die Zählung
    cleaned = re.sub(r"[#\*`_\[\]()>~|]", " ", text)
    return len(cleaned.split())


def is_substantive_md(filepath: str) -> bool:
    """Prüft ob eine .md-Datei substantiell genug für docx-Konvertierung ist.
    
    Kriterien: > 500 Wörter ODER Frontmatter-Tag enthält 'recherche', 'zusammenfassung',
    'analyse', 'report', 'konzept'.
    """
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return False

    if count_words(content) >= 500:
        return True

    # Prüfe Frontmatter-Tags
    fm_match = re.search(r"^---\n(.*?)\n---", content, re.DOTALL)
    if fm_match:
        fm_text = fm_match.group(1).lower()
        substantive_tags = ["recherche", "zusammenfassung", "analyse", "report", "konzept", "deliverable"]
        if any(tag in fm_text for tag in substantive_tags):
            return True

    # Prüfe Dateiname auf substantielle Keywords
    basename = os.path.basename(filepath).lower()
    if any(kw in basename for kw in ["recherche", "zusammenfassung", "analyse", "report", "konzept"]):
        return True

    return False


def convert_md_to_docx(md_path: str) -> str | None:
    """Konvertiert eine .md-Datei nach .docx via pandoc.
    
    Returns der Pfad zur .docx-Datei oder None bei Fehler.
    """
    os.makedirs(DOCX_OUTPUT_DIR, exist_ok=True)

    # Generiere Dateinamen: <WHI-XXX>-<kurz>.docx
    basename = os.path.basename(md_path).replace(".md", "")
    # Kürze auf sinnvolle Länge
    if len(basename) > 60:
        basename = basename[:57] + "..."
    docx_name = f"{basename}.docx"
    docx_path = os.path.join(DOCX_OUTPUT_DIR, docx_name)

    # Wenn .docx schon existiert und jünger als die .md, verwende sie
    try:
        if os.path.exists(docx_path):
            md_mtime = os.path.getmtime(md_path)
            docx_mtime = os.path.getmtime(docx_path)
            if md_mtime <= docx_mtime:
                return docx_path  # schon aktuell
    except OSError:
        pass

    try:
        result = os.popen(
            f'pandoc "{md_path}" -f markdown --wrap=none -o "{docx_path}" 2>&1'
        ).read()
        if os.path.exists(docx_path) and os.path.getsize(docx_path) > 100:
            return docx_path
        else:
            log("WARN", f"pandoc conversion produced empty/small file for {md_path}: {result[:200]}")
            return None
    except Exception as e:
        log("ERROR", f"pandoc conversion failed for {md_path}: {e}")
        return None


# --- Lücke 2: Markdown-Tabelle → .xlsx Rendering ------------------------------

def extract_markdown_tables(text: str) -> list[tuple[str, list[list[str]]]]:
    """Extrahiert alle Markdown-Pipe-Tabellen aus Text.
    
    Returns Liste von (header_row, data_rows) Tupeln.
    """
    tables = []
    lines = text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if "|" in line and not line.startswith("|") is False:
            # Prüfe ob es eine Tabellenzeile ist
            cells = [c.strip() for c in line.split("|") if c.strip()]
            if len(cells) >= 2:
                header = cells
                # Prüfe ob nächste Zeile ein Separator ist (---|---)
                if i + 1 < len(lines):
                    sep = lines[i + 1].strip()
                    if re.match(r"^\|?[-:]+\|?([-:]+\|?)+$", sep):
                        i += 2
                        # Lies Datenzeilen bis Leerzeile oder Ende
                        data = []
                        while i < len(lines):
                            dline = lines[i].strip()
                            if not dline or "|" not in dline:
                                break
                            data.append([c.strip() for c in dline.split("|") if c.strip()])
                            i += 1
                        tables.append((header, data))
                        continue
        i += 1
    return tables


def render_table_to_xlsx(header: list[str], data: list[list[str]], output_path: str) -> bool:
    """Rendert eine Markdown-Tabelle als .xlsx Datei."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabelle"[:31]  # Excel max sheet name length

        # Header
        for col, cell_value in enumerate(header, 1):
            cell = ws.cell(row=1, column=col, value=cell_value)
            cell.font = cell.font.copy(bold=True)

        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Auto-width (grobe Schätzung)
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return True
    except ImportError:
        log("ERROR", "openpyxl not available for xlsx rendering")
        return False
    except Exception as e:
        log("ERROR", f"xlsx rendering failed: {e}")
        return False


def find_or_render_xlsx(md_path: str, revision_number: int = 1) -> list[str]:
    """Findet existierende .xlsx-Datei für eine MD-Tabelle oder rendert neu.
    
    Returns Liste der xlsx-Pfade (kann mehrere sein, wenn mehrere Tabellen im Doc).
    """
    xlsx_files = []

    # Versuche MD zu lesen und Tabellen zu extrahieren
    try:
        with open(md_path, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return []

    tables = extract_markdown_tables(content)
    if not tables:
        return []  # Keine Tabellen, nichts zu tun

    basename = os.path.basename(md_path).replace(".md", "")
    xlsx_dir = os.path.join(os.path.dirname(md_path), "xlsx-exports")
    os.makedirs(xlsx_dir, exist_ok=True)

    for idx, (header, data) in enumerate(tables):
        table_name = f"{basename}-Tabelle{idx+1}" if len(tables) > 1 else basename
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "-", table_name)[:50]
        xlsx_path = os.path.join(xlsx_dir, f"{safe_name}-Rev{revision_number}.xlsx")

        if not os.path.exists(xlsx_path) or render_table_to_xlsx(header, data, xlsx_path):
            if os.path.exists(xlsx_path):
                xlsx_files.append(xlsx_path)

    return xlsx_files





def is_substantive_md(path: str) -> bool:
    """Prüft, ob eine .md-Datei substantiell genug für docx-Konvertierung ist.

    Faustregel: >500 Wörter ODER Frontmatter-Tag enthält
    'recherche', 'zusammenfassung', 'analyse', 'report'.

    Triviale Pfade werden ausgeschlossen: /E-Mails/, /Heartbeat/, /Logs/.
    """
    # Pfad-Ausschluss
    trivial_dirs = ('/E-Mails/', '/Heartbeat/', '/Logs/')
    if any(d in path for d in trivial_dirs):
        return False

    # Wortanzahl prüfen
    try:
        result = subprocess.run(
            ['wc', '-w', path], capture_output=True, text=True, timeout=10
        )
        word_count = int(result.stdout.strip().split()[0])
        if word_count < 500:
            return False
    except (FileNotFoundError, ValueError, subprocess.TimeoutExpired) as e:
        log("WARN", f"word count check failed for {path}: {e}")
        return False

    # Frontmatter-Tags prüfen (nur erste 4KB)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            header = f.read(4096)
        tags_match = re.search(r'tags:\s*\[(.*?)\]', header, re.DOTALL)
        if tags_match:
            tags = [t.strip().strip("'\"") for t in tags_match.group(1).split(',')]
            substantive_tags = {'recherche', 'zusammenfassung', 'analyse', 'report'}
            if any(tag in substantive_tags for tag in tags):
                return True
    except Exception as e:
        log("WARN", f"frontmatter check failed for {path}: {e}")

    return True


def convert_md_to_docx(md_path: str) -> str | None:
    """Konvertiert eine Markdown-Datei zu .docx via pandoc.

    Output-Pfad: gleiche Location wie .md, aber mit .docx Endung.
    Returns den Pfad der erzeugten .docx oder None bei Fehler.
    """
    docx_path = md_path.rsplit('.', 1)[0] + '.docx'

    # Nicht konvertieren, wenn .docx bereits existiert (und jünger ist)
    if os.path.exists(docx_path):
        try:
            md_mtime = os.path.getmtime(md_path)
            docx_mtime = os.path.getmtime(docx_path)
            if md_mtime <= docx_mtime:
                return docx_path  # bereits aktuell
        except OSError:
            pass

    try:
        result = subprocess.run(
            ['pandoc', '-f', 'markdown', '-t', 'docx', '-o', docx_path, md_path],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode != 0:
            log("WARN", f"pandoc conversion failed for {md_path}: {result.stderr[:200]}")
            return None
        if os.path.exists(docx_path):
            log("INFO", f"converted {md_path} -> {docx_path}")
            return docx_path
    except FileNotFoundError:
        log("WARN", "pandoc not found — skipping docx conversion for all files")
    except subprocess.TimeoutExpired:
        log("WARN", f"pandoc timeout for {md_path}")
    except Exception as e:
        log("WARN", f"pandoc error for {md_path}: {e}")
    return None


def select_inline_attachments(vault_files: list[str]) -> tuple[list[str], list[str]]:
    """Returns (inline_files, attachment_files).

    Nur Markdown wird inline in den Body eingebettet — Office-/PDF-Formate
    sind binär und dürfen nicht als Text gelesen werden, daher ausschließlich
    als Anhang. Große .md-Dateien (>MAX_INLINE_FILE_BYTES) gehen ebenfalls nur
    als Anhang raus.

    Lücke 1: Substantielle .md-Dateien werden zusätzlich als .docx konvertiert
    und den Attachments hinzugefügt.
    """
    inline: list[str] = []
    total = 0
    attach: list[str] = []
    for p in vault_files:
        try:
            size = os.path.getsize(p)
        except OSError:
            continue
        attach.append(p)
        if not p.lower().endswith(INLINE_EXT):
            continue
        if size <= MAX_INLINE_FILE_BYTES and total + size <= MAX_ATTACH_TOTAL_BYTES:
            inline.append(p)
            total += size

    # Lücke 1: Substantielle .md → .docx Konvertierung
    for p in vault_files:
        if not p.lower().endswith(".md"):
            continue
        if is_substantive_md(p):
            docx_path = convert_md_to_docx(p)
            if docx_path and os.path.exists(docx_path):
                attach.append(docx_path)

    # Lücke 2: Markdown-Tabellen → .xlsx Rendering (nur Issue-Dokumente im Vault)
    for p in vault_files:
        if not p.lower().endswith(".md"):
            continue
        # Nur Issue-Dokumente (in Paperclip/Projekte oder Paperclip/_Meta)
        if "Paperclip/" not in p:
            continue
        xlsx_files = find_or_render_xlsx(p)
        for xf in xlsx_files:
            if os.path.exists(xf) and xf not in attach:
                attach.append(xf)

    return inline, attach


def _guess_mime_type(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    return {
        ".md": "text/markdown; charset=utf-8",
        ".pdf": "application/pdf",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }.get(ext, "application/octet-stream")


def _build_attachment_payload(paths: list[str]) -> list[dict]:
    """Read each file, base64-encode, return mailhub-conformant attachment objects.

    Mailhub requires {filename, content (base64), mimeType?}. Total payload is
    capped at MAX_ATTACH_TOTAL_BYTES (raw bytes); files past the cap are dropped
    with a warning so the mail still goes through.
    """
    import base64
    out: list[dict] = []
    total = 0
    for p in paths:
        try:
            size = os.path.getsize(p)
        except OSError as exc:
            log("WARN", f"attachment skip {p}: {exc}")
            continue
        if total + size > MAX_ATTACH_TOTAL_BYTES:
            log("WARN", f"attachment cap reached, dropping {os.path.basename(p)} ({size} bytes)")
            continue
        try:
            with open(p, "rb") as f:
                content_b64 = base64.b64encode(f.read()).decode("ascii")
        except OSError as exc:
            log("WARN", f"attachment read failed {p}: {exc}")
            continue
        out.append({
            "filename": os.path.basename(p),
            "content": content_b64,
            "mimeType": _guess_mime_type(p),
        })
        total += size
    return out


def send_mail(subject: str, text_body: str, html_body: str, attachments: list[str]) -> int:
    payload: dict = {
        "from": FROM_ADDR,
        "to": TO_ADDR,
        "subject": subject,
        "text": text_body,
        "attachments": _build_attachment_payload(attachments),
    }
    if html_body:
        payload["html"] = html_body
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        WEBHOOK_URL,
        data=data,
        headers={
            "Content-Type": "application/json",
            "X-Mailhub-Secret": MAILHUB_SECRET,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return r.status
    except urllib.error.HTTPError as e:
        log("ERROR", f"mailhub HTTP {e.code}: {e.read()[:300].decode('utf-8','replace')}")
        return e.code
    except Exception as e:
        log("ERROR", f"mailhub send failed: {e}")
        return 0


# --- Main-Logik ---------------------------------------------------------------


def process_issue(token: str, root: dict, state: dict, dry_run: bool = False) -> bool:
    root_id = root["id"]
    completed_at = root.get("completedAt") or ""
    key = root_id
    if not dry_run and state.get("sent", {}).get(key, {}).get("completedAt") == completed_at:
        return False  # already sent for this completion

    log("INFO", f"processing {root.get('identifier')} (completedAt={completed_at})")

    tree = fetch_full_tree(token, root_id, (root.get("_company") or COMPANIES[0])["id"])
    comments_by_issue = {iid: fetch_comments(token, iid) for iid in tree}
    vault_files = discover_vault_files(tree, comments_by_issue)
    inline_files, attach_files = select_inline_attachments(vault_files)

    subject, text_body, html_body = build_mail(root, tree, comments_by_issue, vault_files, inline_files)

    if dry_run:
        print(f"\n========== DRY-RUN: {root.get('identifier')} ==========")
        print(f"Subject: {subject}")
        print(f"Attachments: {len(attach_files)} files")
        for p in attach_files:
            print(f"  - {p}")
        print(f"HTML body: {len(html_body)} chars (will be sent if non-empty)")
        print("--- Text body (first 4000 chars) ---")
        print(text_body[:4000])
        print("--- End body preview ---")
        return True

    status = send_mail(subject, text_body, html_body, attach_files)
    if status in (200, 201, 202):
        log(
            "SENT",
            f"{root.get('identifier')} tree={len(tree)} docs={len(attach_files)} http={status}",
        )
        state.setdefault("sent", {})[key] = {
            "completedAt": completed_at,
            "sentAt": datetime.now(timezone.utc).isoformat(),
            "identifier": root.get("identifier"),
            "docCount": len(attach_files),
        }
        save_state(state)
        return True
    else:
        log("ERROR", f"send failed for {root.get('identifier')} http={status}")
        return False


def find_candidate_roots(token: str) -> list[dict]:
    out: list[dict] = []
    for company in COMPANIES:
        d = api_get(token, f"/companies/{company['id']}/issues?status=done&limit={LOOKBACK_LIMIT}")
        items = d if isinstance(d, list) else d.get("issues", [])
        for i in items:
            if not isinstance(i, dict):
                continue
            if i.get("parentId") or not i.get("completedAt"):
                continue
            if company["creator_filter"] and i.get("createdByUserId") != company["creator_filter"]:
                continue
            allow = company["identifier_allowlist"]
            if allow is not None and i.get("identifier") not in allow:
                continue
            i["_company"] = company  # für build_mail / fetch_full_tree
            out.append(i)
    return out


def initialize_state_if_first_run(state: dict, candidates: list[dict]) -> bool:
    """On very first run, mark all current done issues as 'already seen' to avoid spam-storm."""
    if state.get("initialized"):
        return False
    log("INFO", f"first run — marking {len(candidates)} existing done issues as seen")
    for c in candidates:
        state.setdefault("sent", {})[c["id"]] = {
            "completedAt": c.get("completedAt"),
            "sentAt": None,
            "identifier": c.get("identifier"),
            "skippedReason": "first_run_baseline",
        }
    state["initialized"] = True
    state["initializedAt"] = datetime.now(timezone.utc).isoformat()
    save_state(state)
    return True


ALERT_THROTTLE_S = 6 * 60 * 60  # höchstens alle 6 h eine Token-Warnmail


def send_auth_alert(state: dict, reason: str) -> None:
    """Warnt Walter per mailhub (token-unabhängig), wenn die API uns abweist.

    Rate-limited auf einmal alle 6 h, damit der 60-s-Poll keine Mailflut auslöst.
    State-Eintrag `authAlert` hält den letzten Sendezeitpunkt fest.
    """
    now = datetime.now(timezone.utc)
    last = state.get("authAlert", {}).get("lastAt")
    if last:
        try:
            if (now - datetime.fromisoformat(last)).total_seconds() < ALERT_THROTTLE_S:
                return  # schon kürzlich gewarnt
        except ValueError:
            pass
    subject = "⚠️ Deliverable-Watcher: Paperclip-Board-Token abgelaufen"
    text = (
        "Der walter-deliverable-watcher kann die Paperclip-API nicht mehr erreichen.\n\n"
        f"Grund: {reason}\n\n"
        "Folge: Es werden KEINE Ergebnis-Mails mehr versendet, bis der Token erneuert ist.\n"
        "Board-Tokens laufen nach 30 Tagen ab — bitte neuen Token via CLI-Auth-Challenge "
        "erzeugen und in ~/.paperclip/auth.json eintragen.\n"
    )
    status = send_mail(subject, text, "", [])
    if status in (200, 201, 202):
        state["authAlert"] = {"lastAt": now.isoformat(), "reason": reason}
        save_state(state)
        log("WARN", f"auth alert mail sent (http={status})")
    else:
        log("ERROR", f"auth alert mail FAILED (http={status})")


def run_once(dry_run_issue: str | None = None) -> int:
    token = load_token()
    state = load_state()
    try:
        candidates = find_candidate_roots(token)
    except AuthError as e:
        log("ERROR", f"AUTH FAILURE — {e}")
        send_auth_alert(state, str(e))
        return 2  # sauberer Exit ohne Stacktrace-Spam; launchd retryt in 60 s

    if dry_run_issue:
        target = next(
            (c for c in candidates if c.get("identifier") == dry_run_issue or c.get("id") == dry_run_issue),
            None,
        )
        if not target:
            log("ERROR", f"dry-run target not found: {dry_run_issue}")
            return 1
        process_issue(token, target, state, dry_run=True)
        return 0

    if initialize_state_if_first_run(state, candidates):
        return 0

    new_sent = 0
    try:
        for c in candidates:
            if process_issue(token, c, state, dry_run=False):
                new_sent += 1
    except AuthError as e:
        log("ERROR", f"AUTH FAILURE mid-run — {e}")
        send_auth_alert(state, str(e))
        return 2
    if new_sent:
        log("INFO", f"sent {new_sent} new deliverable mails")
    return 0


def run_forever() -> None:
    log("INFO", "watcher started")
    while True:
        try:
            run_once()
        except Exception as e:
            log("ERROR", f"run_once crashed: {e}")
        time.sleep(POLL_INTERVAL_S)


def _guarded_once(dry_run_issue: str | None = None) -> int:
    """Fängt unerwartete Fehler ab, damit das err.log nicht mit Stacktraces vollläuft."""
    try:
        return run_once(dry_run_issue=dry_run_issue)
    except Exception as e:  # noqa: BLE001 — bewusst breit: ein Crash darf launchd nicht spammen
        log("ERROR", f"run_once unexpected error: {type(e).__name__}: {e}")
        return 1


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--dry-run":
        target = sys.argv[2] if len(sys.argv) > 2 else None
        sys.exit(_guarded_once(dry_run_issue=target))
    if len(sys.argv) > 1 and sys.argv[1] == "--once":
        sys.exit(_guarded_once())
    if len(sys.argv) > 1 and sys.argv[1] == "--selftest-alert":
        st = load_state()
        send_auth_alert(st, "SELBSTTEST — bitte ignorieren")
        sys.exit(0)
    run_forever()
