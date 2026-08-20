#!/usr/bin/env python3
"""Baut die LLM-Nutzungs-Excel: 3 Sheets + gestapelte Balkengrafiken mit Legende.

Sheet 1  LLM pro Tag        — LLM; Datum; Aufrufe; Token; Laufzeit  + gestapelter Balken (Serie=LLM)
Sheet 2  Agent je Stunde    — Agent; Datum; Stunde; LLM; Aufrufe; Token
Sheet 3  Grafik Agent×LLM   — Matrix + gestapelter Balken (x=Agent, Serie=LLM, Legende)

Usage: build_xlsx.py [TAGE] [AUSGABE.xlsx]
"""
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.legend import Legend
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import pricing
import query

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="1F3864")


def _hms(sec):
    sec = int(sec or 0)
    h, rem = divmod(sec, 3600)
    m, s = divmod(rem, 60)
    return f"{h}:{m:02d}:{s:02d}" if h else f"{m}:{s:02d}"


def _style_header(ws, row, ncols, col0=1):
    for c in range(col0, col0 + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, widths, col0=1):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(col0 + i)].width = w


def _stacked_chart(title, height):
    ch = BarChart()
    ch.type = "col"
    ch.grouping = "stacked"
    ch.overlap = 100
    ch.title = title
    ch.y_axis.title = "Aufrufe"
    ch.height = height
    ch.width = 26
    ch.legend = Legend()
    ch.legend.position = "b"      # Legende mittig unter dem Diagramm
    ch.legend.overlay = False
    return ch


# --------------------------------------------------------------------------- #
def sheet_llm_per_day(wb, days):
    ws = wb.active
    ws.title = "LLM pro Tag"
    ws["A1"] = f"LLM-Nutzung pro Tag (letzte {days} Tage) — Paperclip-Agenten"
    ws["A1"].font = TITLE_FONT
    hdr = ["LLM", "Datum", "Aufrufe", "Input-Token", "Cache-Token", "Output-Token",
           "Token gesamt", "Laufzeit (Sek.)", "Laufzeit (h:m:s)", "Kosten (EUR)"]
    ws.append([])
    ws.append(hdr)
    hrow = ws.max_row
    _style_header(ws, hrow, len(hdr))
    for (model, tag, calls, in_tok, cached_tok, out_tok,
         tokens, dur) in query.per_llm_per_day(days):
        # Kosten je Tag rechnen, damit ein auslaufender Einfuehrungspreis
        # (pricing.INTRO) am richtigen Datum greift.
        kosten = pricing.kosten_eur(model, in_tok, cached_tok, out_tok, tag)
        ws.append([model, str(tag), calls, in_tok or 0, cached_tok or 0,
                   out_tok or 0, tokens or 0, dur or 0, _hms(dur),
                   kosten if kosten is not None else "Preis unbekannt"])
    _autosize(ws, [40, 12, 9, 14, 14, 14, 14, 15, 15, 14])
    for r in range(hrow + 1, ws.max_row + 1):
        for c in (4, 5, 6, 7):
            ws.cell(row=r, column=c).number_format = "#,##0"
        zelle = ws.cell(row=r, column=10)
        if isinstance(zelle.value, float):
            zelle.number_format = '#,##0.00 "€"'
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)

    # ---- Grafik-Matrix (Datum × LLM) rechts, Spalte K ff. ----
    tage, modelle, counts = query.matrix_day_by_model(days)
    gcol = 11  # K
    ws.cell(row=hrow, column=gcol, value="Datum").fill = HEAD_FILL
    ws.cell(row=hrow, column=gcol).font = HEAD_FONT
    for j, m in enumerate(modelle, start=1):
        cell = ws.cell(row=hrow, column=gcol + j, value=m)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
    for i, tag in enumerate(tage, start=1):
        ws.cell(row=hrow + i, column=gcol, value=tag)
        for j, m in enumerate(modelle, start=1):
            ws.cell(row=hrow + i, column=gcol + j, value=counts.get((tag, m), 0))
    ws.column_dimensions[get_column_letter(gcol)].width = 12

    ch = _stacked_chart(f"Aufrufe je LLM pro Tag ({days} Tage)", 11)
    data = Reference(ws, min_col=gcol + 1, max_col=gcol + len(modelle),
                     min_row=hrow, max_row=hrow + len(tage))
    cats = Reference(ws, min_col=gcol, min_row=hrow + 1, max_row=hrow + len(tage))
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, f"A{ws.max_row + 3}")


def sheet_agent_hour(wb, days):
    ws = wb.create_sheet("Agent je Stunde")
    ws["A1"] = f"Agenten-LLM-Aufrufe je Stunde (letzte {days} Tage)"
    ws["A1"].font = TITLE_FONT
    hdr = ["Agent", "Datum", "Stunde", "LLM", "Aufrufe", "Token"]
    ws.append([])
    ws.append(hdr)
    hrow = ws.max_row
    _style_header(ws, hrow, len(hdr))
    for agent, tag, stunde, model, calls, tokens in query.agent_hour(days):
        ws.append([agent, str(tag), stunde, model, calls, tokens or 0])
    _autosize(ws, [24, 12, 9, 40, 9, 12])
    for r in range(hrow + 1, ws.max_row + 1):
        ws.cell(row=r, column=6).number_format = "#,##0"
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)


def sheet_agent_chart(wb, days):
    gs = wb.create_sheet("Grafik Agent×LLM")
    gs["A1"] = f"Aufrufe je Agent, aufgeschlüsselt nach LLM ({days} Tage)"
    gs["A1"].font = TITLE_FONT
    agenten, modelle, agg = query.matrix_agent_by_model(days)
    gs.append([])
    hrow = gs.max_row + 1
    gs.cell(row=hrow, column=1, value="Agent").fill = HEAD_FILL
    gs.cell(row=hrow, column=1).font = HEAD_FONT
    for j, m in enumerate(modelle, start=1):
        cell = gs.cell(row=hrow, column=1 + j, value=m)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
    for i, agent in enumerate(agenten, start=1):
        gs.cell(row=hrow + i, column=1, value=agent)
        for j, m in enumerate(modelle, start=1):
            gs.cell(row=hrow + i, column=1 + j, value=agg.get((agent, m), 0))
    _autosize(gs, [24] + [22] * len(modelle))

    ch = _stacked_chart(f"Aufrufe je Agent nach LLM ({days} Tage)",
                        max(9, 0.42 * len(agenten) + 3))
    data = Reference(gs, min_col=2, max_col=1 + len(modelle),
                     min_row=hrow, max_row=hrow + len(agenten))
    cats = Reference(gs, min_col=1, min_row=hrow + 1, max_row=hrow + len(agenten))
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    anchor_col = get_column_letter(1 + len(modelle) + 2)
    gs.add_chart(ch, f"{anchor_col}{hrow}")


def main():
    days = int(sys.argv[1]) if len(sys.argv) > 1 else 7
    out = sys.argv[2] if len(sys.argv) > 2 else f"LLM-Nutzung-{days}Tage.xlsx"
    wb = Workbook()
    sheet_llm_per_day(wb, days)
    sheet_agent_hour(wb, days)
    sheet_agent_chart(wb, days)
    wb.save(out)
    print(f"geschrieben: {out}")


if __name__ == "__main__":
    main()
